import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
import sys
# путь к корневой директории проекта
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from bot.config import STUDENTS_FILES_DIR, TASKS_FILES_DIR, EXCEL_FILES_DIR
from bot.create_excel import create

#Функция для форматирования столбцов по длине
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 1

# Функция для извлечения даты из имени файла
def extract_date_from_filename(file_name):

    date_pattern = r"tasks_(\d{2}\.\d{2}\.\d{4})"
    match = re.search(date_pattern, file_name)
    if match:
        return match.group(1)
    return None

# Функция для извлечения фио из имени файла
def extract_surname_from_filename(file_name):
    return os.path.splitext(file_name)[0]

def write_to_excel_name(surname, excel_folder_path, output_file="new_wb.xlsx"):
    excel_file_path = os.path.join(excel_folder_path, output_file)

    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    ws['A1'] = "Дата, на которую было дз"
    ws['A2'] = "Номера задач"      

    if ws.max_row == 1 and ws['A1'].value == "Дата, на которую было дз":
        next_row = 3  # Начинаем с третьей строки, после заголовков
    else:
        next_row = ws.max_row + 1

    ws[f'A{next_row}'] = surname

    auto_adjust_column_width(ws)

    wb.save(excel_file_path)

# Функция для записи даты и задач в файл Excel
def write_tasks_to_excel(date, tasks, excel_folder_path, output_file="new_wb.xlsx"):
    excel_file_path = os.path.join(excel_folder_path, output_file)
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Дата")
        ws.cell(row=2, column=1, value="Задачи")

    # Найти первую пустую колонку, чтобы не перезаписывать старые данные
    next_column = ws.max_column + 1 if ws.max_column > 1 else 2

    # Записываем дату в первую строку
    ws.cell(row=1, column=next_column, value=date)

    # Записываем задачи в следующей строке
    for i, task in enumerate(tasks):
        ws.cell(row=2, column=next_column + i, value=task)

    auto_adjust_column_width(ws)

    wb.save(excel_file_path)
    print(f"Дата {date} и задачи добавлены в файл: {excel_file_path}")

 # Функция извлечения фамилии из файла                
def process_files_surname(folder_path_students, excel_folder_path):

    for file_name in os.listdir(folder_path_students):
        surname = extract_surname_from_filename(file_name)
        if surname:
            #print(f"Фамилия извлечена: {surname}")
            write_to_excel_name(surname, excel_folder_path)
        else:
            print(f"В файле {file_name} не найдена фамилия.")

def write_student_task_status(folder_path_students, excel_folder_path, output_file="new_wb.xlsx"):
    excel_file_path = os.path.join(excel_folder_path, output_file)

    if not os.path.exists(excel_file_path):
        print(f"Файл {excel_file_path} не существует. Создайте его сначала.")
        return

    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Получаем список всех задач из второй строки Excel
    task_headers = [ws.cell(row=2, column=col).value for col in range(2, ws.max_column + 1)]
    
    # Обработка файлов с фамилиями и выполненными задачами
    for file_name in os.listdir(folder_path_students):
        file_path = os.path.join(folder_path_students, file_name)
        surname = extract_surname_from_filename(file_name)
        
        if not surname:
            print(f"Фамилия не найдена в файле: {file_name}")
            continue

        # Создаём словарь для задач, которые выполнил ученик
        student_tasks = {}  # {дата: [список задач]}
        
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if ":" in line:  # Обработка строки с задачами
                    date_part, tasks_part = line.split(":", 1)
                    tasks_list = [task.strip() for task in tasks_part.split(",")]
                    student_tasks[date_part.strip()] = tasks_list
        # Найти строку ученика или создать новую
        student_row = None
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == surname:
                student_row = row
                break

        if not student_row:
            student_row = ws.max_row + 1
            ws.cell(row=student_row, column=1, value=surname)

        # Сравнение задач и запись результата
        for col_index, task_number in enumerate(task_headers, start=2):
            if task_number:  # Проверяем, что задача есть в заголовке
                completed = False
                # Сравниваем номер задачи с задачами ученика
                for task_list in student_tasks.values():
                    if str(task_number) in task_list:
                        completed = True
                        break

                ws.cell(row=student_row, column=col_index, value="+" if completed else "-")

    # Автоматическое форматирование ширины столбцов
    auto_adjust_column_width(ws)
    
    # Сохранение изменений
    wb.save(excel_file_path)
    print(f"Статус выполненных задач обновлён в файле: {excel_file_path}")


def add_statistics_row(task_files_dir, excel_folder_path, output_file="new_wb.xlsx"):
    excel_file_path = os.path.join(excel_folder_path, output_file)

    if not os.path.exists(excel_file_path):
        print("Файл Excel не существует.")
        return

    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Найти количество фамилий (последняя строка с фамилиями)
    last_name_row = 3  # Начальная строка с фамилиями
    while ws.cell(row=last_name_row, column=1).value:
        last_name_row += 1
    last_name_row -= 1  # Последняя строка с фамилиями

    # Добавляем строку 'в среднем, %' под всеми фамилиями
    avg_row_plus = last_name_row + 1
    avg_row_minus = avg_row_plus + 1

    ws.cell(row=avg_row_plus, column=1, value="в среднем, %")
    ws.cell(row=avg_row_minus, column=1, value="")  # Пустая ячейка для форматирования

    current_col = 2  # Начинаем со второго столбца

    # Обработка всех файлов задач из папки
    task_files = sorted(os.listdir(task_files_dir))
    for task_file in task_files:
        if task_file.endswith(".txt"):
            file_path = os.path.join(task_files_dir, task_file)
            
            # Считываем количество задач из файла
            with open(file_path, 'r', encoding='utf-8') as f:
                tasks = [line.strip() for line in f if line.strip().isdigit()]
            
            num_tasks = len(tasks)  # Количество задач (колонок в блоке)
            if num_tasks == 0:
                continue  # Пропустить файл, если задач нет
            
            plus_count = 0
            minus_count = 0

            # Подсчёт "+" и "-" по текущему блоку задач
            for col in range(current_col, current_col + num_tasks):
                for row in range(3, last_name_row + 1):  # Для всех фамилий
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value == "+":
                        plus_count += 1
                    elif cell_value == "-":
                        minus_count += 1

            total_tasks = plus_count + minus_count

            # Вычисляем проценты
            if total_tasks > 0:
                plus_percent = round((plus_count / total_tasks) * 100)
                minus_percent = 100 - plus_percent
            else:
                plus_percent = minus_percent = 0

            # Записываем статистику для текущего блока
            ws.cell(row=avg_row_plus, column=current_col, value=f"+  {plus_percent}")
            ws.cell(row=avg_row_minus, column=current_col, value=f"-  {minus_percent}")

            # Переход к следующему блоку задач
            current_col += num_tasks   # Смещаемся вправо на n + 1 столбцов

    # Сохранение изменений
    wb.save(excel_file_path)
    print("Статистика добавлена под каждым блоком задач.")

# Функция статистики задач, которые решили меньше трети учеников
def add_low_completion_tasks(excel_folder_path, output_file="new_wb.xlsx"):
    excel_file_path = os.path.join(excel_folder_path, output_file)

    if not os.path.exists(excel_file_path):
        print("Файл Excel не существует.")
        return

    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Найти количество фамилий (последняя строка с фамилиями)
    last_name_row = 3  # Начальная строка с фамилиями
    while ws.cell(row=last_name_row, column=1).value:
        last_name_row += 1
    last_name_row -= 1  # Последняя строка с фамилиями

    # Строка для новой статистики
    tasks_row_label = last_name_row + 2  # Через одну строку после фамилий
    tasks_row_data = tasks_row_label + 1

    # Добавляем заголовок для задач
    ws.cell(row=tasks_row_label, column=1, value="задачи, которые сделали меньше трети учеников")

    num_students = last_name_row - 2  # Количество учеников

    # Проходим по каждому столбцу начиная со 2-го
    for col in range(2, ws.max_column + 1):
        task_number = ws.cell(row=2, column=col).value  # Номер задачи из строки 2
        plus_count = 0

        # Подсчитываем количество плюсов в столбце
        for row in range(3, last_name_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value == "+":
                plus_count += 1

        # Проверка: если задача решена <= 1/3 учеников
        if plus_count <= num_students / 3:
            # Записываем номер задачи в строку статистики
            ws.cell(row=tasks_row_data, column=col, value=task_number)

    # Сохранение изменений
    wb.save(excel_file_path)
    print("Статистика задач с низким выполнением добавлена.")


def process_files_tasks(folder_path_tasks, excel_folder_path):
    for file_name in sorted(os.listdir(folder_path_tasks)):
        if file_name.startswith("tasks_") and file_name.endswith(".txt"):
            date = extract_date_from_filename(file_name)
            if date:
                file_path = os.path.join(folder_path_tasks, file_name)
                with open(file_path, 'r', encoding='utf-8') as f:
                    tasks = [line.strip() for line in f if line.strip() and line.strip().isdigit()]
                write_tasks_to_excel(date, tasks, excel_folder_path)
            else:
                print(f"В файле {file_name} не найдена дата.")
            
def write():
    try:
        create()
        process_files_surname(STUDENTS_FILES_DIR, EXCEL_FILES_DIR)
        process_files_tasks(TASKS_FILES_DIR, EXCEL_FILES_DIR)
        write_student_task_status(STUDENTS_FILES_DIR, EXCEL_FILES_DIR)
        add_statistics_row(TASKS_FILES_DIR, EXCEL_FILES_DIR)
        add_low_completion_tasks(EXCEL_FILES_DIR)
        print("Статистика успешно сгенерирована.")
    except Exception as e:
        print(f"Ошибка при генерации статистики: {e}")

write()

